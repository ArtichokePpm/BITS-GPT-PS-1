from flask import Flask, jsonify, request
from flask_cors import CORS
from models import *
import uuid
import logging
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

logging.basicConfig(level=logging.DEBUG)

dummy_admins = [
    {'username': 'user1', 'password': 'password1'},
    {'username': 'user2', 'password': 'password2'}
]

@app.route('/api/departments', methods=['GET'])
def get_departments():
    departments_list = []
    for dept in departments_collection.find():
        departments_list.append({
            '_id': str(dept['_id']),
            'dept_name': dept['dept_name'],
            'prog_name': dept['prog_name'],
            'dept_id': dept['dept_id']
        })
    return jsonify({'departments': departments_list})

@app.route('/api/departments', methods=['POST'])
def add_department():
    data = request.json
    new_department = Department(data['dept_name'], data['prog_name'], str(uuid.uuid4()))
    departments_collection.insert_one(new_department.__dict__)
    return jsonify({'message': 'Department added successfully'})

@app.route('/api/courses', methods=['POST'])
def add_course():
    data = request.json
    course_id = str(uuid.uuid4())

    department = departments_collection.find_one({'dept_id': data['department_id']})
    if not department:
        return jsonify({'message': 'Department not found'}), 404

    collection_name = f"{data['course_name']}-{department['dept_name']}-{department['prog_name']}"

    new_course = Course(data['course_name'], data['file_path'], data['department_id'], course_id)
    courses_collection.insert_one(new_course.__dict__)

    knowledge_base = client['Knowledge_base']
    knowledge_base.create_collection(collection_name)
    
    logging.debug(f"Collection created: {collection_name}")

    return jsonify({'message': 'Course added successfully', 'course_id': course_id, 'collection_name': collection_name})

@app.route('/api/upload', methods=['POST'])
def upload_file():
    file = request.files['file']
    course_id = request.form['course_id']
    file_path = f'uploads/{file.filename}'
    file.save(file_path)

    course = courses_collection.find_one({'course_id': course_id})
    if not course:
        return jsonify({'message': 'Course not found'}), 404

    department = departments_collection.find_one({'dept_id': course['department_id']})
    if not department:
        return jsonify({'message': 'Department not found'}), 404

    collection_name = f"{course['course_name']}-{department['dept_name']}-{department['prog_name']}"

    with open(file_path, 'r') as f:
        content = f.read()

    data = []
    current_qa_pair = {}
    for line in content.split('\n'):
        line = line.strip()
        if line.startswith('question:'):
            if current_qa_pair:
                data.append(current_qa_pair)
            current_qa_pair = {"question": line[9:].strip(), "answer": ""}
        elif line.startswith('answer:'):
            current_qa_pair["answer"] = line[7:].strip()

    if current_qa_pair:
        data.append(current_qa_pair)

    knowledge_base = client['Knowledge_base']
    course_collection = knowledge_base[collection_name]
    
    logging.debug(f"Inserting data into collection: {collection_name}")
    logging.debug(f"Data: {data}")

    if data:
        course_collection.insert_many(data)

    return jsonify({'file_path': file_path, 'message': 'File uploaded and processed successfully'})

@app.route('/api/users', methods=['POST'])
def add_users():
    data = request.json
    users_collection.insert_many(data['users'])
    return jsonify({'message': 'Users added successfully'})

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')

    for admin in dummy_admins:
        if admin['username'] == username and admin['password'] == password:
            return jsonify({'message': 'Login successful'}), 200

    return jsonify({'message': 'Invalid credentials'}), 401

@app.route('/api/upload-user-details', methods=['POST'])
def upload_user_details():
    if 'file' not in request.files:
        return jsonify({'message': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No selected file'}), 400

    if file:
        try:
            logging.debug('Processing file: %s', file.filename)
            wb = load_workbook(file)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                user = dict(zip(headers, row))
                # Check for unique User_ID
                if not user_details_collection.find_one({'User_ID': user['User_ID']}):
                    data.append(user)

            logging.debug('Excel data: %s', data)
            if data:
                user_details_collection.insert_many(data)
            return jsonify({'message': 'User details uploaded successfully'}), 200
        except Exception as e:
            logging.error('Error processing file: %s', str(e))
            return jsonify({'message': f'Error processing file: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True)
